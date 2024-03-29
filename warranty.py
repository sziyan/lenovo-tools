#from msilib.schema import Error
import requests
from bs4 import BeautifulSoup
import re
import json
from openpyxl import load_workbook
from datetime import datetime
import time
from telegram import Bot

import logging
import os

try:
    import config
except:
    pass

try:
    token = os.environ.get('TOKEN')
except:
    print('TOKEN is not set!')

bot = Bot(token=token)
logging.basicConfig(level=logging.INFO, filename='output.log', filemode='a', format='%(asctime)s %(levelname)s - %(message)s', datefmt='%d-%b-%y %I:%M:%S %p')


def get_url(model, serial):
    if model.lower() == 'thinkpad e14':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-edge-laptops/thinkpad-e14-type-20ra-20rb/20ra/20ra003ssg/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad e490':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-edge-laptops/thinkpad-e490-type-20n8-20n9/20n8/20n8005psg/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad l13':
        brand, url = 'lenovo', ['https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l13-type-20r3-20r4/20r3/20r3000vsg/{}/warranty'.format(serial), 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l13-type-20r3-20r4/20r4/20r4cto1ww/{}/warranty'.format(serial)]
    elif model.lower() == 'thinkpad l380':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l380-type-20m5-20m6/20m6/20m6cto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad l390':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l390-type-20nr-20ns/20ns/20nscto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad l480':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l480-type-20ls-20lt/20lt/20ltcto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad l490':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l490-type-20q5-20q6/20q6/20q6cto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad x280':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-x-series-laptops/thinkpad-x280-type-20kf-20ke/20ke/20kesbu300/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad x390':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-x-series-laptops/thinkpad-x390/20q1/20q1cto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad l14':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-l-series-laptops/thinkpad-l14-type-20u1-20u2/20u2/20u2cto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad x13':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/sg/en/products/laptops-and-netbooks/thinkpad-x-series-laptops/thinkpad-x13-type-20t2-20t3/20t3/20t3cto1ww/{}/warranty'.format(serial)
    elif model.lower() == 'thinkbook 13s g2':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/us/en/products/laptops-and-netbooks/thinkbook-series/thinkbook-13s-g2-itl/20v9/20v9005ksb/{}/warranty'.format(serial)
    elif model.lower() == 'thinkpad x13 g2':
        brand, url = 'lenovo', 'https://pcsupport.lenovo.com/us/en/products/laptops-and-netbooks/thinkpad-x-series-laptops/thinkpad-x13-yoga-gen-2-type-20wk-20wl/20wl/20wls10p00/{}/warranty'.format(serial)
    elif model.lower() == 'acer':
        brand, url = 'acer', 'http://support.acer.com.sg/support/checkwarrantyresults.asp'
    else:
        brand, url = -1, -1
    return brand, url

def get_warranty(url):
    if isinstance(url, list) is False: #not a list
        warranty = lenovo_bs4(url)
    else:
        for warrantyurl in url: #multiple url provided. Go thru list to make sure warranty url returns valid warranty date
            warranty = lenovo_bs4(warrantyurl)
            if  warranty != -1:
                break
    return warranty          

def lenovo_bs4(url):
    r = requests.get(url)
    data = r.text
    soup = BeautifulSoup(data, 'html.parser')
    script = soup.find_all('script')
    for sr in script:
        try:
            if 'var ds_warranties' in sr.string:
                warranty_string = sr.string
                ds_warranties = re.findall('window.ds_warranties \|\| ({[\w\W]+});', warranty_string)[0]
                js = json.loads(ds_warranties)
                warranty = js.get('BaseUpmaWarranties')[0].get('End')
                break
            else:
                warranty = -1
        except TypeError:
            pass
    return warranty
    

def get_acer_warranty(serial):
    url = 'http://support.acer.com.sg/support/checkwarrantyresults.asp'
    formData = {
        '__VIEWSTATE': 'wEPDwUJNDA4Mzc4NTUxZGQazEtBaHJpnK06W95ZaNAwjHPGdC3C2rmOWwCv9qkuBw==',
        'pserialno': serial
    }

    r = requests.post(url, data = formData)
    data = r.text
    soup = BeautifulSoup(data, 'html.parser')
    row = soup.find_all('td')
    for i in range(len(row)):
        try:
            if row[i].strong.text == 'Onsite Expiry:': #take date of onsite expiry as warranty end date
                warranty = row[i+1].text
                break
        except:
            warranty = -1 #error or page not found
    return warranty

def format_date(brand, date_str):
    try:
        if brand == 'lenovo':
            unformatted = datetime.strptime(date_str,'%Y-%m-%d')
            formatted_date = unformatted.strftime('%d %b %Y')
        else:
            unformatted = datetime.strptime(date_str, '%d %B, %Y')
            formatted_date = unformatted.strftime('%d %b %Y')
        return formatted_date
    except ValueError:
        logging.error('{} does not contain a valid date format'.format(date_str))
        return 'Invalid date format'

def generate_output(row):
    try:
        serial = row[0].value.strip()
        model = row[1].value
        brand, url = get_url(model, serial)
        if url == -1 or brand == -1:
            row[2].value = 'Cannot find model'
            print('{} - Cannot find model'.format(serial))
        elif brand == 'lenovo':
            warranty_date = get_warranty(url)
            output_date = format_date(brand, warranty_date)
            row[2].value = output_date
            print('{} - {}'.format(serial, output_date))
        else: #acer brand
            warranty_date = get_acer_warranty(serial)
            if warranty_date == -1:
                output_date = 'Not Found'
            elif warranty_date.lower() == 'not valid': #warranty page returns Not Valid which most likely is expired
                output_date = 'Expired'
            else:
                output_date = format_date(brand, warranty_date)
            row[2].value = output_date
            print('{} - {}'.format(serial, output_date))
        return 1
    except IndexError as e:
        return -1
    except Exception as e:
        print('Unexpected error: {}'.format(e))
        return -2

def main(chat_id):
    wb = load_workbook(filename='input.xlsx')
    list_of_sheetnames = wb.sheetnames
    #sheet = wb['Sample serial number']
    first_sheet_name = list_of_sheetnames[0] #get the name of 1st sheet
    sheet = wb[first_sheet_name] #set it to sheet variable for openpyxl to work on it
    num_rows = len(sheet['A']) -1 #to factor in top row which serves as title
    logging.info('Received input of {} rows.'.format(num_rows))
    bot.send_message(chat_id=chat_id,text='Rows detected: {}'.format(num_rows))
    noError = 0
    noSuccess = 0
    # set how often a notification of the rows being processed is shown. If more then 300 rows, will notify every 100 rows, else every 50 rows  
    if num_rows > 300:
        process_noti = 100
    else:
        process_noti = 50
    for row in sheet.iter_rows(min_row=2, max_col=3):
        index = int(row[0].row) - 1 # the row that it is currently on. Minus 1 to factor in top row which serves as title
        if row[0].value == '':
            result = 0
        else:
            result = generate_output(row)
        if index % process_noti == 0 : #output message every 50 rows
            bot.send_message(chat_id=chat_id, text='Processing {}/{}'.format(index, num_rows))
        if result == 1:
            noSuccess+=1
        elif result == -1:    #IndexError during function run
            print('{} - IndexError, retrying..'.format(row[0].value))
            time.sleep(5)   #sleep 5 seconds before re-trying function call
            result2 =generate_output(row)   #attempt to re-run function to get warranty
            if result2 == -1:   #if still IndexError, will output to output file
                print('{} - Index Error on 2nd try'.format(row[0].value))
                row[2].value = 'Index Error'
                logging.error('{} index error after 2nd try'.format(row[0].value))
                noError+=1
        elif result == -2:
            print('{} - Unexpected Error'.format(row[0].value))
            row[2].value = 'Unexpected Error'
            logging.error('{} - Unexpected Error'.format(row[0].value))
            noError+=1
        elif result == 0:
            pass
        else:
            print('{} - This should not happen - Invalid error code'.format(row[0].value))
            row[2].value = 'This should not happen - Invalid error code'
            logging.error('{} - This should not happen'.format(row[0].value))

    wb.save('result.xlsx')
    print('Output created..')
    resultSummary = 'Success: {}    Error: {}'.format(noSuccess, noError)
    logging.info(resultSummary)
    bot.send_message(chat_id=chat_id, text = resultSummary)
